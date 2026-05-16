"""
convert_data.py
───────────────
Automatically converts your 3 school Excel files into the 5 CSV files
the Hybrid Scheduler needs.

HOW TO USE:
  1. Place this script in your PrototypeHybridScheduler folder
  2. Create an "input_excel" folder next to it
  3. Put your 3 Excel files inside "input_excel/"
  4. Double-click "convert_data.bat"  (or run: python convert_data.py)
  5. The 5 CSVs will be written to "real_dataset/"

EXPECTED INPUT FILES (filenames can vary, script auto-detects):
  - Faculty schedule  — one sheet per instructor
  - Room utilization  — one sheet per room
  - Class schedule    — one sheet per section
"""

import os
import re
import ast
import sys
import glob
import openpyxl
import pandas as pd
from collections import defaultdict

# ── paths ─────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR   = os.path.join(BASE_DIR, "input_excel")
OUTPUT_DIR  = os.path.join(BASE_DIR, "real_dataset")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── timeslot helpers ──────────────────────────────────────────────────────────
DAYS = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]
DAY_ABBR = {
    "MONDAY":    "Mon", "TUESDAY":  "Tue", "WEDNESDAY": "Wed",
    "THURSDAY":  "Thu", "FRIDAY":   "Fri", "SATURDAY":  "Sat",
}
TIME_ABBR = {
    "6:00 - 7:00":   "0600", "7:00 - 8:00":   "0700", "8:00 - 9:00":   "0800",
    "9:00 - 10:00":  "0900", "10:00 - 11:00": "1000", "11:00 - 12:00": "1100",
    "12:00 - 1:00":  "1200", "1:00 - 2:00":   "1300", "2:00 - 3:00":   "1400",
    "3:00 - 4:00":   "1500", "4:00 - 5:00":   "1600", "5:00 - 6:00":   "1700",
}
TIME_DISPLAY = {
    "0600": "06:00-07:00", "0700": "07:00-08:00", "0800": "08:00-09:00",
    "0900": "09:00-10:00", "1000": "10:00-11:00", "1100": "11:00-12:00",
    "1200": "12:00-13:00", "1300": "13:00-14:00", "1400": "14:00-15:00",
    "1500": "15:00-16:00", "1600": "16:00-17:00", "1700": "17:00-18:00",
}

def normalise_time(raw):
    """Normalise various time formats to 'H:MM - H:MM' key."""
    if not raw:
        return None
    s = re.sub(r"\s+", " ", str(raw).strip())
    # already in our map?
    if s in TIME_ABBR:
        return s
    # try matching pattern like "7:00-8:00" or "7:00 – 8:00"
    m = re.match(r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})", s)
    if m:
        key = f"{m.group(1)} - {m.group(2)}"
        if key in TIME_ABBR:
            return key
    return None

def make_ts_id(day, time_raw):
    t = normalise_time(time_raw)
    if not t:
        return None
    code = TIME_ABBR.get(t)
    abbr = DAY_ABBR.get(day.upper())
    if code and abbr:
        return f"TS{abbr}_{code}"
    return None

# ── file auto-detection ───────────────────────────────────────────────────────
def find_excel_files():
    """Find the 3 Excel files in input_excel/ by inspecting sheet names."""
    xlsx_files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
    if not xlsx_files:
        print(f"\n[ERROR] No .xlsx files found in: {INPUT_DIR}")
        print("        Please create the 'input_excel' folder and place your 3 Excel files inside it.")
        sys.exit(1)

    faculty_file = room_file = class_file = None

    for path in xlsx_files:
        try:
            wb   = openpyxl.load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            wb.close()

            # heuristic: faculty file has many sheets with names like surnames
            # room file has sheets that are mostly numbers/room codes
            # class file has sheets like "CS 1101", "IT 2104"
            section_pattern = re.compile(r'^[A-Z]{2,4}\s*\d{3,4}', re.I)
            room_pattern    = re.compile(r'^\d{3}$|^LAB|^ITL|^Smart|^EDL', re.I)

            section_count = sum(1 for s in sheets if section_pattern.match(s))
            room_count    = sum(1 for s in sheets if room_pattern.match(s))

            fname = os.path.basename(path).lower()

            if "faculty" in fname or "instructor" in fname:
                faculty_file = path
            elif "room" in fname or "utilization" in fname:
                room_file = path
            elif "class" in fname or "schedule" in fname or section_count > 5:
                class_file = path
            elif room_count > 5:
                room_file = path
            elif section_count > 5:
                class_file = path
            else:
                # fallback: assign by process of elimination
                if faculty_file is None:
                    faculty_file = path
                elif room_file is None:
                    room_file = path
                else:
                    class_file = path

        except Exception as e:
            print(f"  [WARN] Could not read {path}: {e}")

    return faculty_file, room_file, class_file

# ── 1. TIMESLOTS ──────────────────────────────────────────────────────────────
def generate_timeslots():
    print("  Generating timeslots.csv ...")
    rows = [("timeslot", "display_time")]
    for day in DAYS:
        abbr = DAY_ABBR[day]
        day_label = day.capitalize()[:3]
        for code, display in TIME_DISPLAY.items():
            rows.append((f"TS{abbr}_{code}", f"{day_label} {display}"))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    out = os.path.join(OUTPUT_DIR, "timeslots.csv")
    df.to_csv(out, index=False)
    print(f"    ✔ {len(df)} timeslots → {out}")
    return df

# ── 2. ROOMS ─────────────────────────────────────────────────────────────────
def convert_rooms(path):
    print(f"  Converting rooms from: {os.path.basename(path)}")
    wb   = openpyxl.load_workbook(path, read_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        caps = []
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            for cell in row:
                try:
                    v = float(str(cell))
                    if 10 < v < 500:
                        caps.append(int(v))
                except:
                    pass
        cap = max(caps) if caps else 40
        rows.append({"room_id": str(sheet_name).strip(), "capacity": cap})
    wb.close()

    df  = pd.DataFrame(rows)
    out = os.path.join(OUTPUT_DIR, "rooms.csv")
    df.to_csv(out, index=False)
    print(f"    ✔ {len(df)} rooms → {out}")
    return df

# ── 3. INSTRUCTORS ────────────────────────────────────────────────────────────
def convert_instructors(path):
    print(f"  Converting instructors from: {os.path.basename(path)}")
    wb   = openpyxl.load_workbook(path, read_only=True)
    rows = []
    inst_name_to_id = {}

    for i, sheet_name in enumerate(wb.sheetnames):
        ws       = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        # find full name
        full_name = str(sheet_name).strip()
        for row in all_rows[:8]:
            if not row:
                continue
            for j, cell in enumerate(row):
                if cell and "name of faculty" in str(cell).lower():
                    for k in range(j + 1, len(row)):
                        if row[k]:
                            full_name = str(row[k]).strip()
                            break

        # find header row containing TIME and day names
        header_row_idx = None
        for ri, row in enumerate(all_rows):
            if not row:
                continue
            row_upper = [str(c).strip().upper() if c else "" for c in row]
            if "TIME" in row_upper and any(d in row_upper for d in DAYS):
                header_row_idx = ri
                break

        if header_row_idx is None:
            # no timetable found — instructor available all slots
            iid = f"I{i+1:03d}"
            rows.append({"instructor_id": iid, "name": full_name,
                         "available_timeslots": "[]"})
            inst_name_to_id[sheet_name.upper()] = iid
            continue

        header    = [str(c).strip().upper() if c else "" for c in all_rows[header_row_idx]]
        day_cols  = {day: header.index(day) for day in DAYS if day in header}

        avail = set()
        for row in all_rows[header_row_idx + 1:]:
            if not row or not row[0]:
                continue
            time_raw = str(row[0]).strip()
            if not re.match(r"\d+:\d+", time_raw):
                continue
            for day, col in day_cols.items():
                if col < len(row) and row[col] and str(row[col]).strip():
                    ts = make_ts_id(day, time_raw)
                    if ts:
                        avail.add(ts)

        iid = f"I{i+1:03d}"
        rows.append({
            "instructor_id": iid,
            "name":          full_name,
            "available_timeslots": str(sorted(avail)),
        })
        last_name = full_name.split(",")[0].strip().upper()
        inst_name_to_id[last_name]          = iid
        inst_name_to_id[sheet_name.upper()] = iid

    wb.close()
    df  = pd.DataFrame(rows)
    out = os.path.join(OUTPUT_DIR, "instructors.csv")
    df.to_csv(out, index=False)
    print(f"    ✔ {len(df)} instructors → {out}")
    return df, inst_name_to_id

# ── 4. COURSES + 5. STUDENTS (sections) ───────────────────────────────────────
def convert_classes(path, inst_name_to_id):
    print(f"  Converting courses & sections from: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path, read_only=True)

    JUNK = {
        "NAME OF ADVISER", "NAME", "PREPARED BY", "REVIEWED BY",
        "CONFORME", "DATE", "IT", "CS", "GED", "SUBJECT",
    }
    BAD_WORDS = [
        "palad", "melo", "gutierrez", "prepared", "reviewed",
        "date", "conforme", "vice", "dean", "dept", "adviser",
        "faculty", "noted", "approved",
    ]

    course_instructor = {}   # course_id -> raw instructor string
    sections          = []

    for sheet_name in wb.sheetnames:
        # skip helper sheets
        if "portal" in sheet_name.lower():
            continue
        sections.append(str(sheet_name).strip())

        ws       = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        sub_idx = inst_idx = None
        for row in all_rows:
            if not row:
                continue
            row_s = [str(c).strip() if c else "" for c in row]
            if "SUBJECT" in [r.upper() for r in row_s]:
                upper = [r.upper() for r in row_s]
                sub_idx  = upper.index("SUBJECT")
                inst_idx = sub_idx + 1 if sub_idx + 1 < len(upper) else None
                continue

            if sub_idx is None:
                continue

            if sub_idx < len(row) and row[sub_idx]:
                cid = str(row[sub_idx]).strip()
                cid_upper = cid.upper()

                # skip junk rows
                if (not cid
                        or cid_upper in JUNK
                        or not re.match(r"^[A-Z]{2,}", cid)
                        or any(b in cid.lower() for b in BAD_WORDS)
                        or len(cid) > 30):
                    continue

                inst = ""
                if inst_idx and inst_idx < len(row) and row[inst_idx]:
                    inst = str(row[inst_idx]).strip()

                if cid not in course_instructor:
                    course_instructor[cid] = inst

    wb.close()

    # match courses to instructor IDs
    course_rows = []
    for cid, inst_name in course_instructor.items():
        iid = "I001"  # fallback
        inst_upper = inst_name.upper()
        for key, vid in inst_name_to_id.items():
            if key and key in inst_upper:
                iid = vid
                break
        course_rows.append({"course_id": cid, "instructor_id": iid, "prerequisite": ""})

    # save courses.csv
    df_courses = pd.DataFrame(course_rows)
    out_courses = os.path.join(OUTPUT_DIR, "courses.csv")
    df_courses.to_csv(out_courses, index=False)
    print(f"    ✔ {len(df_courses)} courses → {out_courses}")

    # save students.csv (sections)
    df_students = pd.DataFrame({"student_id": sections})
    out_students = os.path.join(OUTPUT_DIR, "students.csv")
    df_students.to_csv(out_students, index=False)
    print(f"    ✔ {len(df_students)} sections → {out_students}")

    return df_courses, df_students

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  HYBRID SCHEDULER — DATA CONVERTER")
    print("=" * 60)
    print(f"\n  Input  folder : {INPUT_DIR}")
    print(f"  Output folder : {OUTPUT_DIR}\n")

    # check input folder exists
    if not os.path.isdir(INPUT_DIR):
        os.makedirs(INPUT_DIR)
        print(f"[INFO] Created '{INPUT_DIR}' — please place your 3 Excel files inside it and run again.")
        input("\nPress Enter to exit...")
        sys.exit(0)

    print("  Detecting Excel files...\n")
    faculty_file, room_file, class_file = find_excel_files()

    missing = []
    if not faculty_file: missing.append("Faculty Schedule")
    if not room_file:    missing.append("Room Utilization")
    if not class_file:   missing.append("Class Schedule")

    if missing:
        print(f"\n[ERROR] Could not detect the following files: {', '.join(missing)}")
        print("        Make sure all 3 Excel files are inside the 'input_excel' folder.")
        input("\nPress Enter to exit...")
        sys.exit(1)

    print(f"  Faculty file  : {os.path.basename(faculty_file)}")
    print(f"  Room file     : {os.path.basename(room_file)}")
    print(f"  Class file    : {os.path.basename(class_file)}\n")
    print("-" * 60)

    try:
        generate_timeslots()
        convert_rooms(room_file)
        df_inst, inst_map = convert_instructors(faculty_file)
        convert_classes(class_file, inst_map)

        print("-" * 60)
        print("\n  ✔ All 5 CSV files generated successfully!")
        print(f"  Location: {OUTPUT_DIR}")
        print("\n  You can now run the Hybrid Scheduler.")

    except Exception as e:
        import traceback
        print(f"\n[ERROR] Conversion failed: {e}")
        traceback.print_exc()

    input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()
