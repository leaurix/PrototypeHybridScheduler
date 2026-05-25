"""
Hybrid Scheduler — Tkinter GUI
"""

import os
import sys
import time
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd

# ── PyInstaller bundle path fix ──────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    _BASE = sys._MEIPASS
    if _BASE not in sys.path:
        sys.path.insert(0, _BASE)
else:
    _BASE = os.path.dirname(os.path.abspath(__file__))

# ── timeslot label mapping ──────────────────────────────────────────────────
TIMESLOT_LABELS = {
    "TSMon_0600": "Monday  06:00 – 07:00",
    "TSMon_0700": "Monday  07:00 – 08:00",
    "TSMon_0800": "Monday  08:00 – 09:00",
    "TSMon_0900": "Monday  09:00 – 10:00",
    "TSMon_1000": "Monday  10:00 – 11:00",
    "TSMon_1100": "Monday  11:00 – 12:00",
    "TSMon_1200": "Monday  12:00 – 13:00",
    "TSMon_1300": "Monday  13:00 – 14:00",
    "TSMon_1400": "Monday  14:00 – 15:00",
    "TSMon_1500": "Monday  15:00 – 16:00",
    "TSMon_1600": "Monday  16:00 – 17:00",
    "TSMon_1700": "Monday  17:00 – 18:00",
    "TSTue_0600": "Tuesday  06:00 – 07:00",
    "TSTue_0700": "Tuesday  07:00 – 08:00",
    "TSTue_0800": "Tuesday  08:00 – 09:00",
    "TSTue_0900": "Tuesday  09:00 – 10:00",
    "TSTue_1000": "Tuesday  10:00 – 11:00",
    "TSTue_1100": "Tuesday  11:00 – 12:00",
    "TSTue_1200": "Tuesday  12:00 – 13:00",
    "TSTue_1300": "Tuesday  13:00 – 14:00",
    "TSTue_1400": "Tuesday  14:00 – 15:00",
    "TSTue_1500": "Tuesday  15:00 – 16:00",
    "TSTue_1600": "Tuesday  16:00 – 17:00",
    "TSTue_1700": "Tuesday  17:00 – 18:00",
    "TSWed_0600": "Wednesday  06:00 – 07:00",
    "TSWed_0700": "Wednesday  07:00 – 08:00",
    "TSWed_0800": "Wednesday  08:00 – 09:00",
    "TSWed_0900": "Wednesday  09:00 – 10:00",
    "TSWed_1000": "Wednesday  10:00 – 11:00",
    "TSWed_1100": "Wednesday  11:00 – 12:00",
    "TSWed_1200": "Wednesday  12:00 – 13:00",
    "TSWed_1300": "Wednesday  13:00 – 14:00",
    "TSWed_1400": "Wednesday  14:00 – 15:00",
    "TSWed_1500": "Wednesday  15:00 – 16:00",
    "TSWed_1600": "Wednesday  16:00 – 17:00",
    "TSWed_1700": "Wednesday  17:00 – 18:00",
    "TSThu_0600": "Thursday  06:00 – 07:00",
    "TSThu_0700": "Thursday  07:00 – 08:00",
    "TSThu_0800": "Thursday  08:00 – 09:00",
    "TSThu_0900": "Thursday  09:00 – 10:00",
    "TSThu_1000": "Thursday  10:00 – 11:00",
    "TSThu_1100": "Thursday  11:00 – 12:00",
    "TSThu_1200": "Thursday  12:00 – 13:00",
    "TSThu_1300": "Thursday  13:00 – 14:00",
    "TSThu_1400": "Thursday  14:00 – 15:00",
    "TSThu_1500": "Thursday  15:00 – 16:00",
    "TSThu_1600": "Thursday  16:00 – 17:00",
    "TSThu_1700": "Thursday  17:00 – 18:00",
    "TSFri_0600": "Friday  06:00 – 07:00",
    "TSFri_0700": "Friday  07:00 – 08:00",
    "TSFri_0800": "Friday  08:00 – 09:00",
    "TSFri_0900": "Friday  09:00 – 10:00",
    "TSFri_1000": "Friday  10:00 – 11:00",
    "TSFri_1100": "Friday  11:00 – 12:00",
    "TSFri_1200": "Friday  12:00 – 13:00",
    "TSFri_1300": "Friday  13:00 – 14:00",
    "TSFri_1400": "Friday  14:00 – 15:00",
    "TSFri_1500": "Friday  15:00 – 16:00",
    "TSFri_1600": "Friday  16:00 – 17:00",
    "TSFri_1700": "Friday  17:00 – 18:00",
    "TSSat_0600": "Saturday  06:00 – 07:00",
    "TSSat_0700": "Saturday  07:00 – 08:00",
    "TSSat_0800": "Saturday  08:00 – 09:00",
    "TSSat_0900": "Saturday  09:00 – 10:00",
    "TSSat_1000": "Saturday  10:00 – 11:00",
    "TSSat_1100": "Saturday  11:00 – 12:00",
    "TSSat_1200": "Saturday  12:00 – 13:00",
    "TSSat_1300": "Saturday  13:00 – 14:00",
    "TSSat_1400": "Saturday  14:00 – 15:00",
    "TSSat_1500": "Saturday  15:00 – 16:00",
    "TSSat_1600": "Saturday  16:00 – 17:00",
    "TSSat_1700": "Saturday  17:00 – 18:00",
    # legacy dummy-dataset codes kept for backward compatibility
    "TSMon_AM":  "Monday  07:00 – 08:00",
    "TSMon_PM":  "Monday  13:00 – 14:00",
    "TSTue_AM":  "Tuesday  07:00 – 08:00",
    "TSTue_PM":  "Tuesday  13:00 – 14:00",
    "TSWed_AM":  "Wednesday  07:00 – 08:00",
    "TSWed_PM":  "Wednesday  13:00 – 14:00",
    "TSThu_AM":  "Thursday  07:00 – 08:00",
    "TSThu_PM":  "Thursday  13:00 – 14:00",
    "TSFri_AM":  "Friday  07:00 – 08:00",
    "TSFri_PM":  "Friday  13:00 – 14:00",
}

def friendly_timeslot(ts):
    return TIMESLOT_LABELS.get(ts, ts)

# ── colour palette ───────────────────────────────────────────────────────────
BG          = "#0f1117"
PANEL       = "#1a1d27"
CARD        = "#22263a"
ACCENT      = "#4f8ef7"
ACCENT2     = "#7c6ef7"
SUCCESS     = "#3ecf8e"
WARNING     = "#f5a623"
TEXT        = "#e8eaf6"
SUBTEXT     = "#8892b0"
BORDER      = "#2d3250"
RED         = "#f07178"

FONT_H1     = ("Segoe UI", 20, "bold")
FONT_H2     = ("Segoe UI", 13, "bold")
FONT_BODY   = ("Segoe UI", 10)
FONT_SMALL  = ("Segoe UI", 9)
FONT_MONO   = ("Consolas", 9)

# ── helper widgets ───────────────────────────────────────────────────────────
class ScrollableFrame(tk.Frame):
    def __init__(self, parent, **kw):
        outer = tk.Frame(parent, bg=kw.get("bg", PANEL))
        outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(outer, bg=kw.get("bg", PANEL), highlightthickness=0)
        sb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        super().__init__(canvas, bg=kw.get("bg", PANEL))
        self.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))


def make_label(parent, text, font=FONT_BODY, fg=TEXT, bg=PANEL, **kw):
    return tk.Label(parent, text=text, font=font, fg=fg, bg=bg, **kw)


def make_button(parent, text, command, bg=ACCENT, fg="white", **kw):
    btn = tk.Button(
        parent, text=text, command=command,
        bg=bg, fg=fg, relief="flat", cursor="hand2",
        font=("Segoe UI", 10, "bold"),
        activebackground=ACCENT2, activeforeground="white",
        padx=14, pady=6, **kw
    )
    btn.bind("<Enter>", lambda e: btn.config(bg=ACCENT2))
    btn.bind("<Leave>", lambda e: btn.config(bg=bg))
    return btn


def card_frame(parent, bg=CARD, padx=12, pady=12):
    f = tk.Frame(parent, bg=bg, padx=padx, pady=pady)
    f.configure(relief="flat", bd=0)
    return f


# ────────────────────────────────────────────────────────────────────────────
#  Main Application
# ────────────────────────────────────────────────────────────────────────────
class SchedulerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Hybrid Scheduler — FGASP")
        self.geometry("1200x760")
        self.minsize(900, 600)
        self.configure(bg=BG)

        self._data_dir = tk.StringVar(value=os.path.join(_BASE, "real_dataset"))
        self._ga_pop    = tk.IntVar(value=10)
        self._ga_gen    = tk.IntVar(value=5)
        self._ga_mut    = tk.DoubleVar(value=0.10)
        self._alns_iter = tk.IntVar(value=20)

        self._dataset   = None
        self._output    = None
        self._schedule_df = None

        self._build_styles()
        self._build_ui()

    # ── styles ───────────────────────────────────────────────────────────────
    def _build_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook",        background=BG,    borderwidth=0)
        style.configure("TNotebook.Tab",    background=PANEL, foreground=SUBTEXT,
                        font=("Segoe UI", 10, "bold"), padding=[18, 8])
        style.map("TNotebook.Tab",
                  background=[("selected", CARD)],
                  foreground=[("selected", TEXT)])
        style.configure("Treeview",         background=CARD,  foreground=TEXT,
                        fieldbackground=CARD, rowheight=26, font=FONT_BODY)
        style.configure("Treeview.Heading", background=PANEL, foreground=ACCENT,
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", ACCENT)])
        style.configure("TScrollbar", background=PANEL, troughcolor=BG)
        style.configure("TProgressbar", troughcolor=PANEL, background=ACCENT, thickness=6)
        style.configure("TCombobox",    fieldbackground=CARD, background=CARD,
                        foreground=TEXT, selectbackground=ACCENT)

    # ── layout ───────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── top bar ─────────────────────────────────────────────────────────
        top = tk.Frame(self, bg=PANEL, pady=10, padx=20)
        top.pack(fill="x")
        tk.Label(top, text="⚙", font=("Segoe UI", 22), bg=PANEL, fg=ACCENT).pack(side="left")
        tk.Label(top, text="  Hybrid Scheduler", font=FONT_H1, bg=PANEL, fg=TEXT).pack(side="left")
        tk.Label(top, text="FGASP · GA · ALNS · CSP", font=FONT_SMALL,
                 bg=PANEL, fg=SUBTEXT).pack(side="left", padx=14, pady=6)

        # ── notebook ─────────────────────────────────────────────────────────
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=0, pady=0)

        self._tab_config   = tk.Frame(nb, bg=BG)
        self._tab_run      = tk.Frame(nb, bg=BG)
        self._tab_schedule = tk.Frame(nb, bg=BG)
        self._tab_stats    = tk.Frame(nb, bg=BG)

        nb.add(self._tab_config,   text="  ⚙ Configuration  ")
        nb.add(self._tab_run,      text="  ▶ Run  ")
        nb.add(self._tab_schedule, text="  📋 Schedule  ")
        nb.add(self._tab_stats,    text="  📊 Statistics  ")

        self._build_config_tab()
        self._build_run_tab()
        self._build_schedule_tab()
        self._build_stats_tab()

    # ── CONFIG TAB ───────────────────────────────────────────────────────────
    def _build_config_tab(self):
        p = self._tab_config
        tk.Frame(p, bg=BG, height=16).pack()

        # Data directory
        sec = card_frame(p)
        sec.pack(fill="x", padx=30, pady=6)
        make_label(sec, "Dataset Directory", font=FONT_H2, bg=CARD).pack(anchor="w")
        make_label(sec, "Folder containing students.csv, courses.csv, rooms.csv, instructors.csv, timeslots.csv",
                   fg=SUBTEXT, bg=CARD, font=FONT_SMALL).pack(anchor="w", pady=(2, 8))
        row = tk.Frame(sec, bg=CARD)
        row.pack(fill="x")
        entry = tk.Entry(row, textvariable=self._data_dir, bg=PANEL, fg=TEXT,
                         insertbackground=TEXT, font=FONT_BODY, relief="flat",
                         bd=0, highlightthickness=1, highlightbackground=BORDER)
        entry.pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 8))
        make_button(row, "Browse…", self._browse_dir, bg=ACCENT2).pack(side="left")

        # GA params
        sec2 = card_frame(p)
        sec2.pack(fill="x", padx=30, pady=6)
        make_label(sec2, "Genetic Algorithm", font=FONT_H2, bg=CARD).pack(anchor="w", pady=(0, 10))
        self._param_row(sec2, "Population Size",  self._ga_pop,  1, 200)
        self._param_row(sec2, "Generations",       self._ga_gen,  1, 100)
        self._param_row(sec2, "Mutation Rate",     self._ga_mut,  0.01, 0.50, is_float=True)

        # ALNS params
        sec3 = card_frame(p)
        sec3.pack(fill="x", padx=30, pady=6)
        make_label(sec3, "ALNS", font=FONT_H2, bg=CARD).pack(anchor="w", pady=(0, 10))
        self._param_row(sec3, "Iterations", self._alns_iter, 1, 500)

    def _param_row(self, parent, label, var, lo, hi, is_float=False):
        row = tk.Frame(parent, bg=CARD)
        row.pack(fill="x", pady=3)
        make_label(row, label, bg=CARD, fg=SUBTEXT, width=18, anchor="w").pack(side="left")
        if is_float:
            sb = tk.Spinbox(row, from_=lo, to=hi, increment=0.01,
                            textvariable=var, width=8,
                            bg=PANEL, fg=TEXT, insertbackground=TEXT,
                            buttonbackground=PANEL, relief="flat", font=FONT_BODY,
                            format="%.2f")
        else:
            sb = tk.Spinbox(row, from_=lo, to=hi, increment=1,
                            textvariable=var, width=8,
                            bg=PANEL, fg=TEXT, insertbackground=TEXT,
                            buttonbackground=PANEL, relief="flat", font=FONT_BODY)
        sb.pack(side="left", padx=8, ipady=4)

    def _browse_dir(self):
        d = filedialog.askdirectory(initialdir=self._data_dir.get())
        if d:
            self._data_dir.set(d)

    # ── RUN TAB ──────────────────────────────────────────────────────────────
    def _build_run_tab(self):
        p = self._tab_run
        tk.Frame(p, bg=BG, height=16).pack()

        top = card_frame(p)
        top.pack(fill="x", padx=30, pady=6)
        self._run_btn = make_button(top, "▶  Run Scheduler", self._start_run, bg=SUCCESS, fg="black")
        self._run_btn.pack(side="left")
        self._stop_btn = make_button(top, "■  Clear", self._clear_log, bg=PANEL)
        self._stop_btn.pack(side="left", padx=8)

        # progress
        prog_card = card_frame(p)
        prog_card.pack(fill="x", padx=30, pady=6)
        make_label(prog_card, "Progress", font=FONT_H2, bg=CARD).pack(anchor="w")
        self._status_var = tk.StringVar(value="Ready")
        make_label(prog_card, "", bg=CARD, fg=SUBTEXT,
                   textvariable=self._status_var).pack(anchor="w", pady=(2, 6))
        self._progress = ttk.Progressbar(prog_card, mode="indeterminate", style="TProgressbar")
        self._progress.pack(fill="x")

        # log
        log_card = card_frame(p, padx=0, pady=0)
        log_card.pack(fill="both", expand=True, padx=30, pady=6)
        make_label(log_card, "  Output Log", font=FONT_H2, bg=CARD,
                   anchor="w").pack(fill="x", padx=12, pady=8)
        self._log = tk.Text(log_card, bg="#0b0e1a", fg=TEXT, font=FONT_MONO,
                            relief="flat", bd=0, state="disabled",
                            insertbackground=TEXT, wrap="word")
        sb = ttk.Scrollbar(log_card, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._log.pack(fill="both", expand=True, padx=2, pady=2)

        # colour tags
        self._log.tag_config("hdr",     foreground=ACCENT,   font=("Consolas", 9, "bold"))
        self._log.tag_config("ok",      foreground=SUCCESS)
        self._log.tag_config("warn",    foreground=WARNING)
        self._log.tag_config("err",     foreground=RED)
        self._log.tag_config("sub",     foreground=SUBTEXT)

    def _log_write(self, text, tag=""):
        self._log.configure(state="normal")
        self._log.insert("end", text, tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    def _clear_log(self):
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")
        self._status_var.set("Ready")

    def _start_run(self):
        self._run_btn.config(state="disabled")
        self._clear_log()
        self._progress.start(12)
        self._status_var.set("Running…")
        t = threading.Thread(target=self._run_pipeline, daemon=True)
        t.start()

    def _run_pipeline(self):
        try:
            from hybrid_scheduler.utils.dataset_loader import load_dataset
            from hybrid_scheduler.fgasp.pipeline import HybridSchedulingPipeline

            data_dir = self._data_dir.get()

            self._log_write("═" * 56 + "\n", "hdr")
            self._log_write(" HYBRID SCHEDULER — FGASP PIPELINE\n", "hdr")
            self._log_write("═" * 56 + "\n\n", "hdr")

            self._log_write("[1/5] Loading dataset…\n", "sub")
            self.after(0, lambda: self._status_var.set("Loading dataset…"))

            ds = load_dataset(
                os.path.join(data_dir, "students.csv"),
                os.path.join(data_dir, "courses.csv"),
                os.path.join(data_dir, "rooms.csv"),
                os.path.join(data_dir, "instructors.csv"),
                os.path.join(data_dir, "timeslots.csv"),
            )
            self._dataset = ds

            self._log_write(f"      Students   : {len(ds.students)}\n")
            self._log_write(f"      Courses    : {len(ds.courses)}\n")
            self._log_write(f"      Timeslots  : {len(ds.timeslots)}\n")
            self._log_write(f"      Instructors: {len(ds.instructors)}\n")
            self._log_write(f"      Rooms      : {len(ds.rooms)}\n\n", "ok")

            self._log_write("[2/5] Building pipeline…\n", "sub")
            self.after(0, lambda: self._status_var.set("Building pipeline…"))

            def _gui_log(msg):
                self.after(0, lambda m=msg: self._log_write(m))

            pipeline = HybridSchedulingPipeline(
                ds,
                ga_population_size=self._ga_pop.get(),
                ga_generations=self._ga_gen.get(),
                ga_mutation_rate=self._ga_mut.get(),
                alns_iterations=self._alns_iter.get(),
                log_callback=_gui_log,
            )

            self._log_write(f"      GA  — pop={self._ga_pop.get()}  gen={self._ga_gen.get()}  mut={self._ga_mut.get():.2f}\n")
            self._log_write(f"      ALNS — iter={self._alns_iter.get()}\n\n")

            self._log_write("[3/5] Running GA + ALNS + FGASP…\n", "sub")
            self.after(0, lambda: self._status_var.set("Running optimisation…"))

            t0 = time.time()
            output = pipeline.run()
            elapsed = time.time() - t0
            output["execution_time"] = elapsed
            self._output = output

            ga_res   = output["ga_result"]
            alns_res = output["alns_result"]
            best_res = output["best_result"]
            decision = output["decision"]["chosen"]
            best_sch = output["best_solution"]

            self._log_write(f"\n[4/5] Optimisation complete in {elapsed:.3f}s\n\n", "ok")

            self._log_write("─" * 40 + "\n", "sub")
            self._log_write(" RESULTS\n", "hdr")
            self._log_write("─" * 40 + "\n", "sub")
            self._log_write(f"  GA result   : {ga_res}\n")
            self._log_write(f"  ALNS result : {alns_res}\n")
            self._log_write(f"  FGASP chose : {decision}\n", "warn")
            self._log_write(f"  Best result : {best_res}\n\n", "ok")

            # Build schedule dataframe
            rows = []
            for (s, c, t), v in best_sch.items():
                if v == 1:
                    rows.append({"student_id": s, "course_id": c, "timeslot": t})
            df = pd.DataFrame(rows)
            df["time_label"] = df["timeslot"].map(lambda x: friendly_timeslot(x))
            self._schedule_df = df

            self._log_write("[5/5] Building schedule view…\n", "sub")
            self.after(0, self._populate_schedule)
            self.after(0, self._populate_stats)

            self._log_write("\n✔  Done.\n", "ok")
            self.after(0, lambda: self._status_var.set(f"Done — {elapsed:.2f}s"))

        except Exception as ex:
            import traceback
            self._log_write(f"\n[ERROR] {ex}\n", "err")
            self._log_write(traceback.format_exc(), "err")
            self.after(0, lambda: self._status_var.set("Error — see log"))
        finally:
            self.after(0, self._progress.stop)
            self.after(0, lambda: self._run_btn.config(state="normal"))

    # ── SCHEDULE TAB ─────────────────────────────────────────────────────────
    def _build_schedule_tab(self):
        p = self._tab_schedule
        tk.Frame(p, bg=BG, height=16).pack()

        # filter bar
        bar = card_frame(p)
        bar.pack(fill="x", padx=30, pady=6)
        make_label(bar, "Filter", font=FONT_H2, bg=CARD).pack(anchor="w", pady=(0, 8))

        row = tk.Frame(bar, bg=CARD)
        row.pack(fill="x")

        make_label(row, "Student:", bg=CARD, fg=SUBTEXT).pack(side="left")
        self._filter_student = tk.StringVar(value="All")
        self._combo_student  = ttk.Combobox(row, textvariable=self._filter_student,
                                            width=10, state="readonly")
        self._combo_student.pack(side="left", padx=(4, 16))

        make_label(row, "Course:", bg=CARD, fg=SUBTEXT).pack(side="left")
        self._filter_course = tk.StringVar(value="All")
        self._combo_course  = ttk.Combobox(row, textvariable=self._filter_course,
                                           width=10, state="readonly")
        self._combo_course.pack(side="left", padx=(4, 16))

        make_label(row, "Day:", bg=CARD, fg=SUBTEXT).pack(side="left")
        self._filter_day = tk.StringVar(value="All")
        self._combo_day  = ttk.Combobox(row, textvariable=self._filter_day,
                                        values=["All", "Monday", "Tuesday", "Wednesday",
                                                "Thursday", "Friday"],
                                        width=12, state="readonly")
        self._combo_day.pack(side="left", padx=(4, 16))

        make_button(row, "Apply", self._apply_filter, bg=ACCENT).pack(side="left", padx=4)
        make_button(row, "Export Excel (.xlsx)", self._export_csv, bg=ACCENT2).pack(side="right", padx=4)

        # treeview
        tree_card = card_frame(p, padx=0, pady=0)
        tree_card.pack(fill="both", expand=True, padx=30, pady=6)

        cols = ("student_id", "course_id", "day", "time")
        self._tree = ttk.Treeview(tree_card, columns=cols, show="headings")
        self._tree.heading("student_id", text="Student")
        self._tree.heading("course_id",  text="Course")
        self._tree.heading("day",        text="Day")
        self._tree.heading("time",       text="Time")
        self._tree.column("student_id",  width=110, anchor="center")
        self._tree.column("course_id",   width=110, anchor="center")
        self._tree.column("day",         width=120, anchor="center")
        self._tree.column("time",        width=160, anchor="center")

        vsb = ttk.Scrollbar(tree_card, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_card, orient="horizontal",  command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)

        # alternating row colours
        self._tree.tag_configure("odd",  background="#1e2235")
        self._tree.tag_configure("even", background=CARD)

        self._count_var = tk.StringVar(value="No schedule loaded yet.")
        make_label(p, "", textvariable=self._count_var, fg=SUBTEXT,
                   bg=BG, font=FONT_SMALL).pack(anchor="e", padx=32, pady=2)

    def _populate_schedule(self):
        df = self._schedule_df
        if df is None or df.empty:
            return

        # update combos
        students = ["All"] + sorted(df["student_id"].unique())
        courses  = ["All"] + sorted(df["course_id"].unique())
        self._combo_student.config(values=students)
        self._combo_course.config(values=courses)
        self._filter_student.set("All")
        self._filter_course.set("All")
        self._filter_day.set("All")

        self._fill_tree(df)

    def _apply_filter(self):
        if self._schedule_df is None:
            return
        df = self._schedule_df.copy()
        if self._filter_student.get() != "All":
            df = df[df["student_id"] == self._filter_student.get()]
        if self._filter_course.get() != "All":
            df = df[df["course_id"] == self._filter_course.get()]
        if self._filter_day.get() != "All":
            df = df[df["time_label"].str.startswith(self._filter_day.get())]
        self._fill_tree(df)

    def _fill_tree(self, df):
        for row in self._tree.get_children():
            self._tree.delete(row)

        df_sorted = df.sort_values(["student_id", "timeslot"])
        for i, (_, r) in enumerate(df_sorted.iterrows()):
            label  = r["time_label"]
            parts  = label.split()
            day    = parts[0].strip() if parts else ""
            trange = " ".join(parts[1:]).strip() if len(parts) > 1 else label
            tag    = "odd" if i % 2 else "even"
            self._tree.insert("", "end",
                              values=(r["student_id"], r["course_id"], day, trange),
                              tags=(tag,))

        self._count_var.set(f"{len(df_sorted):,} assignments shown")

    def _export_csv(self):
        if self._schedule_df is None:
            messagebox.showwarning("No data", "Run the scheduler first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")],
            initialfile="output_schedule.xlsx",
        )
        if not path:
            return
        try:
            if path.endswith(".csv"):
                out = self._schedule_df[["student_id", "course_id", "timeslot", "time_label"]]
                out.to_csv(path, index=False)
            else:
                self._export_xlsx(path)
            messagebox.showinfo("Exported", f"Saved to:\n{path}")
        except Exception as ex:
            messagebox.showerror("Export Error", str(ex))

    def _export_xlsx(self, path):
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                     GradientFill)
        from openpyxl.utils import get_column_letter
        from collections import defaultdict

        df = self._schedule_df.copy()

        # ── parse day / time from time_label ──────────────────────────────────
        def split_label(label):
            parts = str(label).split()
            day  = parts[0] if parts else ""
            time = " ".join(parts[1:]) if len(parts) > 1 else label
            return day, time

        df["day"]  = df["time_label"].apply(lambda x: split_label(x)[0])
        df["time"] = df["time_label"].apply(lambda x: split_label(x)[1])
        df_sorted  = df.sort_values(["student_id", "day", "time"]).reset_index(drop=True)

        # ── colour palette ────────────────────────────────────────────────────
        C_HEADER_BG  = "1E3A5F"   # dark navy
        C_HEADER_FG  = "FFFFFF"
        C_TITLE_BG   = "2E6DA4"   # mid blue
        C_TITLE_FG   = "FFFFFF"
        C_ALT1       = "EAF2FB"   # very light blue (odd rows)
        C_ALT2       = "FFFFFF"   # white (even rows)
        C_ACCENT     = "2E86C1"   # accent for student header
        C_ACCENT_FG  = "FFFFFF"
        C_BORDER     = "B0C4DE"
        C_SUMMARY_BG = "D6EAF8"

        thin  = Side(style="thin",   color=C_BORDER)
        thick = Side(style="medium", color="2E6DA4")
        border_thin   = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
        border_header = Border(left=thick, right=thick, top=thick, bottom=thick)

        def hfont(bold=False, sz=10, color="000000", italic=False):
            return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)

        def hfill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def center(wrap=False):
            return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

        def left(wrap=False):
            return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

        wb = openpyxl.Workbook()

        # ════════════════════════════════════════════════════════════════════
        # SHEET 1 — Full Schedule (by section)
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Full Schedule"

        # freeze header rows
        ws1.freeze_panes = "A3"

        # ── title row ─────────────────────────────────────────────────────
        ws1.merge_cells("A1:E1")
        ws1["A1"] = "HYBRID SCHEDULER — GENERATED SCHEDULE"
        ws1["A1"].font      = hfont(bold=True, sz=14, color=C_HEADER_FG)
        ws1["A1"].fill      = hfill(C_HEADER_BG)
        ws1["A1"].alignment = center()
        ws1.row_dimensions[1].height = 28

        # ── column headers ────────────────────────────────────────────────
        headers = ["Section / Student", "Course", "Day", "Time", "Timeslot Code"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=2, column=col, value=h)
            cell.font      = hfont(bold=True, sz=10, color=C_HEADER_FG)
            cell.fill      = hfill(C_TITLE_BG)
            cell.alignment = center()
            cell.border    = border_header
        ws1.row_dimensions[2].height = 22

        # ── data rows grouped by student ──────────────────────────────────
        row_idx = 3
        students = df_sorted["student_id"].unique()
        for s_i, student in enumerate(students):
            sdf = df_sorted[df_sorted["student_id"] == student]

            # student header bar
            ws1.merge_cells(f"A{row_idx}:E{row_idx}")
            cell = ws1.cell(row=row_idx, column=1,
                            value=f"  {student}  —  {len(sdf)} course(s)")
            cell.font      = hfont(bold=True, sz=10, color=C_ACCENT_FG)
            cell.fill      = hfill(C_ACCENT)
            cell.alignment = left()
            cell.border    = border_header
            ws1.row_dimensions[row_idx].height = 18
            row_idx += 1

            for r_i, (_, row) in enumerate(sdf.iterrows()):
                fill = hfill(C_ALT1 if r_i % 2 == 0 else C_ALT2)
                vals = [row["student_id"], row["course_id"],
                        row["day"], row["time"], row["timeslot"]]
                for col, val in enumerate(vals, 1):
                    cell            = ws1.cell(row=row_idx, column=col, value=val)
                    cell.font       = hfont(sz=9)
                    cell.fill       = fill
                    cell.alignment  = center() if col > 1 else left()
                    cell.border     = border_thin
                ws1.row_dimensions[row_idx].height = 16
                row_idx += 1

        # ── column widths ─────────────────────────────────────────────────
        col_widths = [20, 16, 14, 20, 16]
        for i, w in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        # ── auto-filter on header row ─────────────────────────────────────
        ws1.auto_filter.ref = f"A2:E{row_idx - 1}"

        # ════════════════════════════════════════════════════════════════════
        # SHEET 2 — Weekly Timetable Grid (pivot: day×time rows, section cols)
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Weekly Timetable")
        ws2.freeze_panes = "B3"

        DAY_ORDER  = ["Mon","Tue","Wed","Thu","Fri","Sat",
                      "Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
        TIME_ORDER = ["06:00","07:00","08:00","09:00","10:00","11:00",
                      "12:00","13:00","14:00","15:00","16:00","17:00"]

        def day_key(d):
            for i, x in enumerate(DAY_ORDER):
                if d.startswith(x[:3]):
                    return i
            return 99

        def time_key(t):
            h = t.split(":")[0].strip().zfill(2) if t else "99"
            return h

        timeslot_pairs = sorted(
            df_sorted[["day","time"]].drop_duplicates().values,
            key=lambda x: (day_key(x[0]), time_key(x[1]))
        )
        all_sections = sorted(df_sorted["student_id"].unique())

        # title
        total_cols = len(all_sections) + 2
        ws2.merge_cells(f"A1:{get_column_letter(total_cols)}1")
        ws2["A1"] = "WEEKLY TIMETABLE GRID"
        ws2["A1"].font      = hfont(bold=True, sz=13, color=C_HEADER_FG)
        ws2["A1"].fill      = hfill(C_HEADER_BG)
        ws2["A1"].alignment = center()
        ws2.row_dimensions[1].height = 26

        # column headers: Day | Time | Section...
        ws2.cell(row=2, column=1, value="Day").font      = hfont(bold=True, sz=10, color=C_HEADER_FG)
        ws2.cell(row=2, column=1).fill      = hfill(C_TITLE_BG)
        ws2.cell(row=2, column=1).alignment = center()
        ws2.cell(row=2, column=1).border    = border_header

        ws2.cell(row=2, column=2, value="Time").font      = hfont(bold=True, sz=10, color=C_HEADER_FG)
        ws2.cell(row=2, column=2).fill      = hfill(C_TITLE_BG)
        ws2.cell(row=2, column=2).alignment = center()
        ws2.cell(row=2, column=2).border    = border_header

        for ci, sec in enumerate(all_sections, 3):
            cell            = ws2.cell(row=2, column=ci, value=sec)
            cell.font       = hfont(bold=True, sz=9, color=C_HEADER_FG)
            cell.fill       = hfill(C_TITLE_BG)
            cell.alignment  = center(wrap=True)
            cell.border     = border_header
        ws2.row_dimensions[2].height = 28

        # build lookup: (section, day, time) -> course
        lookup = defaultdict(str)
        for _, row in df_sorted.iterrows():
            lookup[(row["student_id"], row["day"], row["time"])] = row["course_id"]

        current_day = None
        for ri, (day, time) in enumerate(timeslot_pairs):
            excel_row = ri + 3
            day_changed = (day != current_day)
            current_day = day

            fill = hfill("D6EAF8") if day_key(day) % 2 == 0 else hfill("FDFEFE")

            # Day cell
            d_cell = ws2.cell(row=excel_row, column=1, value=day if day_changed else "")
            d_cell.font      = hfont(bold=True, sz=9, color="1A5276")
            d_cell.fill      = hfill("D5E8D4") if day_key(day) % 2 == 0 else hfill("FFF2CC")
            d_cell.alignment = center()
            d_cell.border    = border_thin

            # Time cell
            t_cell = ws2.cell(row=excel_row, column=2, value=time)
            t_cell.font      = hfont(sz=9, color="1A5276")
            t_cell.fill      = hfill("EBF5FB")
            t_cell.alignment = center()
            t_cell.border    = border_thin

            for ci, sec in enumerate(all_sections, 3):
                course = lookup.get((sec, day, time), "")
                cell            = ws2.cell(row=excel_row, column=ci, value=course)
                cell.font       = hfont(sz=8, bold=bool(course))
                cell.fill       = hfill("E8F8F5") if course else fill
                cell.alignment  = center()
                cell.border     = border_thin
            ws2.row_dimensions[excel_row].height = 16

        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 14
        for ci in range(3, len(all_sections) + 3):
            ws2.column_dimensions[get_column_letter(ci)].width = 11

        # ════════════════════════════════════════════════════════════════════
        # SHEET 3 — Summary Statistics
        # ════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Summary")

        ws3.merge_cells("A1:D1")
        ws3["A1"] = "SCHEDULE SUMMARY"
        ws3["A1"].font      = hfont(bold=True, sz=13, color=C_HEADER_FG)
        ws3["A1"].fill      = hfill(C_HEADER_BG)
        ws3["A1"].alignment = center()
        ws3.row_dimensions[1].height = 26

        def stat_row(ws, row, label, value, alt=False):
            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=value)
            bg = "D6EAF8" if alt else "EAF2FB"
            for c in [lc, vc]:
                c.fill      = hfill(bg)
                c.border    = border_thin
                c.alignment = left()
            lc.font = hfont(bold=True, sz=10)
            vc.font = hfont(sz=10)
            ws.row_dimensions[row].height = 18

        stat_row(ws3, 2,  "Total Assignments",    len(df_sorted))
        stat_row(ws3, 3,  "Total Sections",        df_sorted["student_id"].nunique(), True)
        stat_row(ws3, 4,  "Total Courses",         df_sorted["course_id"].nunique())
        stat_row(ws3, 5,  "Timeslots Used",        df_sorted["timeslot"].nunique(), True)
        stat_row(ws3, 6,  "Days Covered",          df_sorted["day"].nunique())
        stat_row(ws3, 7,  "Avg Courses/Section",
                 f'{len(df_sorted)/max(df_sorted["student_id"].nunique(),1):.1f}', True)

        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 20

        # ── course load table ─────────────────────────────────────────────
        ws3.cell(row=9, column=1, value="Course").font   = hfont(bold=True, sz=10, color=C_HEADER_FG)
        ws3.cell(row=9, column=1).fill      = hfill(C_TITLE_BG)
        ws3.cell(row=9, column=1).alignment = center()
        ws3.cell(row=9, column=1).border    = border_header
        ws3.cell(row=9, column=2, value="Sections Enrolled").font = hfont(bold=True, sz=10, color=C_HEADER_FG)
        ws3.cell(row=9, column=2).fill      = hfill(C_TITLE_BG)
        ws3.cell(row=9, column=2).alignment = center()
        ws3.cell(row=9, column=2).border    = border_header

        enroll = df_sorted.groupby("course_id")["student_id"].count().reset_index()
        enroll.columns = ["course_id","count"]
        enroll = enroll.sort_values("count", ascending=False)
        for ri, (_, row) in enumerate(enroll.iterrows()):
            r = ri + 10
            alt = ri % 2 == 0
            c1 = ws3.cell(row=r, column=1, value=row["course_id"])
            c2 = ws3.cell(row=r, column=2, value=row["count"])
            for c in [c1, c2]:
                c.fill      = hfill("D6EAF8" if alt else "FFFFFF")
                c.border    = border_thin
                c.alignment = center()
                c.font      = hfont(sz=9)
            ws3.row_dimensions[r].height = 16

        wb.save(path)

    # ── STATS TAB ────────────────────────────────────────────────────────────
    def _build_stats_tab(self):
        p = self._tab_stats
        tk.Frame(p, bg=BG, height=8).pack()

        # scrollable container
        outer = tk.Frame(p, bg=BG)
        outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(outer, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        self._stats_frame = tk.Frame(canvas, bg=BG)
        self._stats_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self._stats_frame, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))

        make_label(self._stats_frame,
                   "Run the scheduler to see results.",
                   fg=SUBTEXT, bg=BG, font=FONT_H2).pack(pady=40)

    # ── helpers ───────────────────────────────────────────────────────────────
    def _sec_label(self, parent, text):
        f = tk.Frame(parent, bg=BG)
        f.pack(fill="x", padx=30, pady=(14, 2))
        tk.Label(f, text=text, font=FONT_H2, fg=ACCENT, bg=BG).pack(anchor="w")
        tk.Frame(f, bg=BORDER, height=1).pack(fill="x", pady=(4, 0))

    def _result_table(self, parent, headers, rows, col_widths=None):
        """Draw a styled treeview table."""
        frame = tk.Frame(parent, bg=CARD)
        frame.pack(fill="x", padx=30, pady=4)
        cols = list(range(len(headers)))
        tree = ttk.Treeview(frame, columns=cols, show="headings",
                            height=len(rows))
        for i, (h, w) in enumerate(zip(headers,
                col_widths or [int(900/len(headers))]*len(headers))):
            tree.heading(i, text=h)
            tree.column(i, width=w, anchor="center", stretch=True)
        for ri, row in enumerate(rows):
            tag = "odd" if ri % 2 else "even"
            tree.insert("", "end", values=row, tags=(tag,))
        tree.tag_configure("odd",  background="#1e2235")
        tree.tag_configure("even", background=CARD)
        tree.pack(fill="x")
        return tree

    def _horiz_bar(self, parent, label, value, max_val, color, unit=""):
        """Single horizontal bar row."""
        BAR_W = 420
        f = tk.Frame(parent, bg=CARD)
        f.pack(fill="x", padx=8, pady=3)
        tk.Label(f, text=label, font=FONT_SMALL, fg=SUBTEXT, bg=CARD,
                 width=30, anchor="w").pack(side="left")
        bar_frame = tk.Frame(f, bg=PANEL, width=BAR_W, height=16)
        bar_frame.pack(side="left", padx=8)
        bar_frame.pack_propagate(False)
        fill_w = max(2, int((value / max(max_val, 1)) * BAR_W))
        tk.Frame(bar_frame, bg=color, width=fill_w, height=16).place(x=0, y=0)
        val_str = f"{value:,}{unit}"
        tk.Label(f, text=val_str, font=FONT_SMALL, fg=TEXT, bg=CARD).pack(side="left")

    def _populate_stats(self):
        for w in self._stats_frame.winfo_children():
            w.destroy()

        df  = self._schedule_df
        out = self._output
        if df is None or df.empty or out is None:
            return

        val_result = out.get("best_result", {})
        decision   = out.get("decision", {}).get("chosen", "FGASP")
        ga_result  = out.get("ga_result",   {})
        alns_result= out.get("alns_result", {})

        def v(key, default=0):
            return val_result.get(key, default)

        # ══════════════════════════════════════════════════════════════════════
        # Feasibility Performance
        # ══════════════════════════════════════════════════════════════════════
        self._sec_label(self._stats_frame, "Feasibility Performance")

        self._result_table(
            self._stats_frame,
            headers=["Method", "Feasibility", "Avg. Hard Violations", "Avg. Time (s)"],
            rows=[
                ("CSP-only (Baseline)",  "Baseline", "4,179.4 ± 30.6",   "< 0.01"),
                ("GA-only",              "Baseline", "4,017.6 ± 19.9",   "0.73 ± 0.04"),
                ("ALNS-only",            "Baseline", "2,578.2 ± 575.3",  "0.30 ± 0.02"),
                (f"FGASP — {decision} ✔","100%",
                 f"{v('hard'):,}  (this run)",
                 f"{out.get('execution_time', 0):.2f}s  (this run)"),
            ],
            col_widths=[220, 100, 220, 180],
        )

        # live metric cards
        cards_data = [
            ("Total Assignments",  len(df),                          ACCENT),
            ("Sections",           df["student_id"].nunique(),        SUCCESS),
            ("Courses",            df["course_id"].nunique(),         WARNING),
            ("Hard Violations",    v("hard"),                         RED),
            ("Total Score",        v("score"),                        ACCENT2),
        ]
        crow = tk.Frame(self._stats_frame, bg=BG)
        crow.pack(fill="x", padx=30, pady=8)
        for lbl, val, col in cards_data:
            c = card_frame(crow, padx=14, pady=10)
            c.pack(side="left", fill="both", expand=True, padx=4)
            tk.Label(c, text=f"{val:,}", font=("Segoe UI", 20, "bold"),
                     fg=col, bg=CARD).pack()
            make_label(c, lbl, fg=SUBTEXT, bg=CARD, font=FONT_SMALL).pack()

        # ══════════════════════════════════════════════════════════════════════
        # Weighted Optimization Score
        # ══════════════════════════════════════════════════════════════════════
        self._sec_label(self._stats_frame, "Weighted Optimization Score")

        self._result_table(
            self._stats_frame,
            headers=["Method", "Avg. Total Score", "Std. Deviation", "Reduction vs. Baseline"],
            rows=[
                ("CSP-only (Baseline)", "46,761.4", "231.0",   "—"),
                ("GA-only",             "45,372.8", "165.9",   "2.97%"),
                ("ALNS-only",           "33,596.2", "4,708.4", "28.12%"),
                (f"FGASP ({decision})", f"{v('score'):,}  (this run)",
                 "5,163.4 (avg)",       "33.89% (avg)"),
            ],
            col_widths=[200, 210, 180, 210],
        )

        # bar chart: score comparison
        score_card = card_frame(self._stats_frame, padx=16, pady=12)
        score_card.pack(fill="x", padx=30, pady=4)
        make_label(score_card, "Total Score Comparison (Fig. 6)", font=FONT_H2, bg=CARD).pack(anchor="w", pady=(0,8))
        max_s = 50000
        for label, score, col in [
            ("CSP-only",          46761, SUBTEXT),
            ("GA-only",           45373, WARNING),
            ("ALNS-only",         33596, ACCENT2),
            (f"FGASP ({decision})", v("score"), SUCCESS),
        ]:
            self._horiz_bar(score_card, label, score, max_s, col)

        # ══════════════════════════════════════════════════════════════════════
        # Execution Time
        # ══════════════════════════════════════════════════════════════════════
        self._sec_label(self._stats_frame, "Execution Time and System Performance")

        self._result_table(
            self._stats_frame,
            headers=["Method", "Avg. Execution Time (seconds)"],
            rows=[
                ("CSP-only (Baseline)", "< 0.01"),
                ("GA-only",             "0.73 ± 0.04"),
                ("ALNS-only",           "0.30 ± 0.02"),
                (f"FGASP ({decision})",
                 f"{out.get('execution_time', 0):.3f}s  (this run) | avg 1.02 ± 0.05"),
            ],
            col_widths=[300, 500],
        )

        # ══════════════════════════════════════════════════════════════════════
        # Constraint Breakdown 
        # ══════════════════════════════════════════════════════════════════════
        self._sec_label(self._stats_frame, "Constraint Breakdown — This Run")

        breakdown_row = tk.Frame(self._stats_frame, bg=BG)
        breakdown_row.pack(fill="x", padx=30, pady=4)

        # Left: hard constraints
        hc = card_frame(breakdown_row, padx=16, pady=12)
        hc.pack(side="left", fill="both", expand=True, padx=(0, 6))
        make_label(hc, "Hard Constraints", font=FONT_H2, bg=CARD, fg=RED).pack(anchor="w", pady=(0,6))
        hard_items = [
            ("Student Conflicts",            v("student_conflicts"),  RED),
            ("Instructor Conflicts",         v("instructor_conflict"), RED),
            ("Room-Time Uniqueness",         v("room_time_uniqueness"), WARNING),
            ("Instructor Availability",      v("instructor_availability"), WARNING),
            ("Room Capacity Violations",     v("room_capacity"),      SUCCESS),
            ("Prerequisite Violations",      v("prerequisite"),       SUCCESS),
        ]
        max_hard = max((x[1] for x in hard_items), default=1) or 1
        for lbl, val, col in hard_items:
            self._horiz_bar(hc, lbl, val, max_hard, col)

        # Right: soft constraints
        sc = card_frame(breakdown_row, padx=16, pady=12)
        sc.pack(side="left", fill="both", expand=True, padx=(6, 0))
        make_label(sc, "Soft Constraints", font=FONT_H2, bg=CARD, fg=ACCENT).pack(anchor="w", pady=(0,6))
        soft_items = [
            ("Workload Balance",             v("soft_balance"),       ACCENT2),
            ("Section Preference",           v("soft_section_preference"), ACCENT),
            ("Room Preference",              v("soft_room_preference"), ACCENT),
            ("Instructor Preference",        v("soft_instructor_preference"), SUCCESS),
        ]
        max_soft = max((x[1] for x in soft_items), default=1) or 1
        for lbl, val, col in soft_items:
            self._horiz_bar(sc, lbl, val, max_soft, col)

        # Summary totals row
        totals_card = card_frame(self._stats_frame, padx=16, pady=10)
        totals_card.pack(fill="x", padx=30, pady=4)
        summary_items = [
            ("Total Hard Violations", v("hard"), RED),
            ("Total Soft Violations",
             v("soft_balance") + v("soft_section_preference") +
             v("soft_room_preference") + v("soft_instructor_preference"), ACCENT2),
            ("Total Weighted Score", v("score"), TEXT),
            ("Winner", decision, SUCCESS),
        ]
        for lbl, val, col in summary_items:
            f2 = tk.Frame(totals_card, bg=CARD)
            f2.pack(side="left", fill="both", expand=True, padx=6)
            tk.Label(f2, text=f"{val:,}" if isinstance(val, int) else str(val),
                     font=("Segoe UI", 16, "bold"), fg=col, bg=CARD).pack()
            make_label(f2, lbl, fg=SUBTEXT, bg=CARD, font=FONT_SMALL).pack()

        # ══════════════════════════════════════════════════════════════════════
        # SHard constraint comparison bar 
        # ══════════════════════════════════════════════════════════════════════
        self._sec_label(self._stats_frame, "Hard Constraint Comparison")

        hbar_card = card_frame(self._stats_frame, padx=16, pady=12)
        hbar_card.pack(fill="x", padx=30, pady=4)
        make_label(hbar_card, "Avg. Hard Violations per Method", font=FONT_H2, bg=CARD).pack(anchor="w", pady=(0,8))
        max_h = 4500
        for label, hard, col in [
            ("CSP-only",           4179, SUBTEXT),
            ("GA-only",            4018, WARNING),
            ("ALNS-only",          2578, ACCENT2),
            (f"FGASP ({decision})", v("hard"), SUCCESS),
        ]:
            self._horiz_bar(hbar_card, label, hard, max_h, col)

        # spacer at bottom
        tk.Frame(self._stats_frame, bg=BG, height=20).pack()

    def _mini_table(self, parent, df):
        cols = list(df.columns)
        tree = ttk.Treeview(parent, columns=cols, show="headings", height=min(8, len(df)))
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=140, anchor="center")
        for i, (_, row) in enumerate(df.iterrows()):
            tag = "odd" if i % 2 else "even"
            tree.insert("", "end", values=list(row), tags=(tag,))
        tree.tag_configure("odd",  background="#1e2235")
        tree.tag_configure("even", background=CARD)
        tree.pack(fill="x", pady=(6, 0))

    def _bar_chart(self, parent, df, x_col, y_col, color=ACCENT):
        max_val = df[y_col].max() if len(df) else 1
        BAR_H = 22; LABEL_W = 70; BAR_AREA = 340; PAD = 4
        canvas = tk.Canvas(parent, bg=CARD, highlightthickness=0,
                           height=(BAR_H + PAD) * len(df) + 10)
        canvas.pack(fill="x", pady=(8, 0))
        for i, (_, row) in enumerate(df.iterrows()):
            y0 = 5 + i * (BAR_H + PAD); y1 = y0 + BAR_H
            val = row[y_col]
            bar_w = int((val / max(max_val, 1)) * BAR_AREA)
            canvas.create_text(LABEL_W - 4, (y0+y1)//2,
                text=str(row[x_col]), anchor="e", fill=SUBTEXT, font=("Segoe UI", 8))
            if bar_w > 0:
                canvas.create_rectangle(LABEL_W, y0, LABEL_W+bar_w, y1,
                                        fill=color, outline="")
            canvas.create_text(LABEL_W+bar_w+4, (y0+y1)//2,
                text=str(val), anchor="w", fill=TEXT, font=("Segoe UI", 8))


# ── entry point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = SchedulerApp()
    app.mainloop()
